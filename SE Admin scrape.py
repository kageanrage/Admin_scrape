import bs4,re, openpyxl, os, sqlite3, requests, time, smtplib, pprint
from openpyxl.styles import Font, NamedStyle, PatternFill
from selenium import webdriver
from bs4 import BeautifulSoup
from config import Config   # this imports the config file where the private data sits

import logging
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.CRITICAL)     # switches off logging

logging.debug('Imported modules')

logging.debug('Start of program')
logging.debug('Checking if Laptop or Desktop (and opening relevant local HTML files if using test HTML)')

cfg = Config()      # create an instance of the Config class, essentially brings private config data into play

old_site_regex = re.compile(
'<a href="https://data.studentedge.com.au/admin/survey/details/(.{36})">(.{1,75})<\/a><\/td><td class="clickable">(.{1,70}?)<\/td><td class="clickable">(.{1,10})<\/td><td class="clickable">(.{1,30})<\/td>(.{80,180})201\d<\/td><td class="clickable">(.{1,10})?<\/td><td class="clickable">(\d+)?<\/td><td class="clickable">(\d+)?<\/td><td class="clickable">(\d+)?<\/td><td class="clickable">(\d+)?<\/td><td class="published t-center clickable"><span class="((True)|(False))">((True)|(False))<\/span><\/td><\/tr><tr class="gridrow(_alternate)? selectable-row"><td class="clickable">')  # doesn't cater for abbreviated URL prefix (changed 26-8 on site)
new_site_regex = re.compile(
'<a href="/admin/survey/details/(.{36})">(.{1,75})<\/a><\/td><td class="clickable">(.{1,70}?)<\/td><td class="clickable">(.{1,10})<\/td><td class="clickable">(.{1,30})<\/td>(.{80,200})201\d<\/td><td class="clickable">(.{1,10})?<\/td><td class="clickable">(\d+)?<\/td><td class="clickable">(\d+)?<\/td><td class="clickable">(\d+)?<\/td><td class="clickable">(\d+)?<\/td><td class="published t-center clickable"><span class="((True)|(False))">((True)|(False))<\/span><\/td><\/tr><tr class="gridrow(_alternate)? selectable-row"><td class="clickable">')  # caters for abbreviated URL prefix (changed 26-8 on site)


# changes logic depending on if I'm using Laptop or Desktop
# Example files - using saved HTML in 2 different directories. Toggle on for test mode, or off for live.

if os.getcwd() == cfg.laptop_dir:   # Using Laptop
    logging.debug('Laptop PC detected')
    T1_file = open(cfg.laptop_T1)
    T1_soup = bs4.BeautifulSoup(T1_file, "html.parser")  # turns the HTML into a beautiful soup object
    T2_file = open(cfg.laptop_T2)
    T2_soup = bs4.BeautifulSoup(T2_file, "html.parser")  # turns the HTML into a beautiful soup object
    T3_file = open(cfg.laptop_T3)
    T3_soup = bs4.BeautifulSoup(T3_file, "html.parser")  # turns the HTML into a beautiful soup object

elif os.getcwd() == cfg.desktop_dir:    # Using Desktop
    logging.debug('Desktop PC detected')
    T1_file = open(cfg.desktop_T1)
    T1_soup = bs4.BeautifulSoup(T1_file, "html.parser")  # turns the HTML into a beautiful soup object
    T2_file = open(cfg.desktop_T2)
    T2_soup = bs4.BeautifulSoup(T2_file, "html.parser")  # turns the HTML into a beautiful soup object
    T3_file = open(cfg.desktop_T3)
    T3_soup = bs4.BeautifulSoup(T3_file, "html.parser")  # turns the HTML into a beautiful soup object


def download_soup():
    chrome_path = r'C:\Program Files\Python37\chromedriver.exe'
    driver = webdriver.Chrome(chrome_path)
    driver.get(cfg.survey_admin_URL)  # load survey admin page
    email_elem = driver.find_element_by_id('UserName') #enter username & password and submit
    email_elem.send_keys(cfg.uname)
    pass_elem = driver.find_element_by_id('Password')
    pass_elem.send_keys(cfg.pwd)
    pass_elem.submit()
    time.sleep(5)   # wait 5 seconds to log in
    content = driver.page_source
    soup = bs4.BeautifulSoup(content, "html.parser")
    # logging.debug('Newly downloaded soup looks like this:\n\n', soup)
    #soupFile = open(htmlFileName, "w")
    #soupFile.write(str(soup))
    #soupFile.close()
    return soup


def process_soup(soup, string_txt_filename, regex):
    logging.debug('Starting table isolation')
    table_only = soup.select(
        'table')  # isolates the table (which is the only bit I need) from the HTML. Type is list, was expecting BS4 object
    #logging.debug('table_only looks like this:\n\n\n',table_only)
    logging.debug('Converting bs4 object into string')
    table_string = str(table_only)  # converts the bs4 object to a string
    logging.debug('writing table_string_file.txt')
    table_string_file = open(string_txt_filename, 'w')
    table_string_file.write(table_string)
    table_string_file.close()
    # logging.debug('table_string looks like this:\n\n\n',table_string)
    # May not be able to isolate further within BS4 so switching to regex to parse.
    # TO DO: create a regex to identify each project on the Admin page

    # TO DO: Return all examples of regex findall search
    logging.debug('Conducting regex findall search')
    mo = regex.findall(table_string)
    #print('newly created mo looks like this:\n\n',mo)
    return mo


def process_string(string_txt_filename, regex):
    table_string_file = open(string_txt_filename)
    table_string_content = table_string_file.read()
    logging.debug('Conducting regex findall search')
    mo = regex.findall(table_string_content)
    print('newly created mo looks like this:\n\n',mo)
    return mo


def list_creator(valueList):   #this function takes in a MO from the regex and creates and returns a per-project list, ordered as per the headings list below
    #headings = ['URL','Alias','Survey name','Project number','Client name','junk','Expected LOI','Actual LOI','Completes','Screen Outs','Quota Fulls','Live on site'] #here I've added 'Live on Site'
    new_list = []
    #logging.debug('Start of list creation for',valueList[3])
    for i in range(0,12):
        new_list.append(valueList[i])
    completes = int(valueList[8])
    quota_fulls = int(valueList[10])
    screen_outs = int(valueList[9])
    if completes == 0 | screen_outs == 0 | quota_fulls == 0:
        incidence = 0
        qf_incidence = 0
    else:
        incidence = (completes / (completes + screen_outs))
        qf_incidence = (completes / (completes + screen_outs + quota_fulls))
    new_list.append(incidence)
    new_list.append(qf_incidence)
    #logging.debug('new_list is:',new_list)
    #logging.debug('valueList is',valueList[0:12])
    #logging.debug('{} C / {} C + {} screen_outs + {} quota_fulls = {} IR.'.format(completes,completes,screen_outs,quota_fulls,incidence))
    #logging.debug(newDict)
    return new_list


def create_masterList(mo):     #creates a list of all projects in given MO, first row will be headings
    #global masterList
    master_list = [['URL', 'Survey name', 'Alias', 'Project number', 'Client name', 'junk', 'Expected LOI', 'Actual LOI',
                   'Completes', 'Screen Outs', 'Quota Fulls', 'Live on site', 'Incidence Rate', 'QF IR']]
    for i in range(0, len(mo) - 1):
        master_list.append(list_creator(mo[i]))
    return master_list


def create_top_list(mo, num):    #num = how long you want the list to be
    top_list = []
    for i in range(0, num):
        top_list.append(list_creator(mo[i]))
    return top_list


def new_project_search(new_list, old_list):

    matches = []
    unmatched = []

    for new_project in new_list:
        unmatched.append(new_project[3])   #this should make a list with all the Project numbers in new_list

    for new_project in new_list:
        for old_project in old_list:
            if new_project[3] == old_project[3]:
                matches.append(new_project[3])
                #if new_project[3] not in unmatched:
                #    raise Exception('Project not found in unmatched list, cannot remove')
                try:
                    unmatched.remove(new_project[3]) #this should remove all matches so that unmatched is the list of non-matched jobs
                except:
                    print(new_project[3],'could not be removed')
                    pass

    #print('Unmatched are as follows: ',unmatched)
    print('List of matched items: ', matches)
    return(unmatched)


def excel_export(list):     #### THIS FUNCTION IS THE EXPORT TO EXCEL  #####
    logging.debug('Excel section - creating workbook object')
    wb = openpyxl.Workbook()  # create excel workbook object
    wb.save('admin.xlsx')  # save workbook as admin.xlsx
    sheet = wb.active  # create sheet object as the Active sheet from the workbook object
    wb.save('admin.xlsx')  # save workbook as admin.xlsx
    # LIST-BASED POPULATION OF EXCEL SHEET
    for row, row_data in enumerate(list,
                                  1):  # where row is a number starting with 1, increasing each loop, and row_data = each masterList item
        for column in range(1, 15):  # where column is a number starting with 1 and ending with 14
            cell = sheet.cell(row=row, column=column)  # so on first loop, row = 2, col = 1
            v = row_data[column - 1]
            try:
                v = float(v)  # try to convert value to a float, so it will store numbers as numbers and not strings
            except ValueError:
                pass  # if it's not a number and therefore returns an error, don't try to convert it to a number
            cell.value = v  # write the value (v) to the cell
            if (column == 13) | (column == 14):  # for all cells in column 13 or 14 (IR / QFIR)
                cell.style = 'Percent'  # ... change cell format (style) to 'Percent', a built-in style within openpyxl

    # this section populates the first row in the sheet (headings) with bold style
    #make_bold(sheet, wb, sheet['A1':'N1'])    #Calls the make_bold function on first row of excel sheet
    wb.save('admin.xlsx')  # save workbook as admin.xlsx
    logging.debug('Excel workbook completed and saved')


def make_bold(sheet, wb, sheet_slice):
    highlight = NamedStyle(name='highlight')
    highlight.font = Font(bold=True)
    wb.add_named_style(highlight)
    for row in sheet_slice:  # iterate over rows in slice (seems redundant as only 1 row but apparently necessary)
        for cell in row:  # iterate over cells in row
            sheet[cell.coordinate].style = highlight  # add bold to each cell


def export_to_sqlite(list_of_projects): # Export to SQLite
    global conn, c
    logging.debug('Initiating SQLite section')
    conn = sqlite3.connect('admin.db')  # define connection - database is created also
    c = conn.cursor()  # define cursor

    def create_table():
        c.execute(
            'CREATE TABLE IF NOT EXISTS surveysTable(URL TEXT, SurveyName TEXT, Alias TEXT, ProjectNumber TEXT, ClientName TEXT, junk TEXT, ExpectedLOI REAL, ActualLOI REAL, Completes REAL, ScreenOuts REAL, QuotaFulls REAL, LiveOnSite TEXT, IncidenceRate REAL, QFIR REAL)')  # creates the table. CAPS for pure SQL, regular casing otherwise.

    def dynamic_data_entry(
            list):  # at the moment if I pass it an ordered list, it will assign that list to the headings. If I convert dictionariesList into a list of lists, this will be simple.
        # Trying to do a lot on this next line, something is up with it
        c.execute(
            "INSERT INTO surveysTable (URL, SurveyName, Alias, ProjectNumber, ClientName, junk, ExpectedLOI, ActualLOI, Completes, ScreenOuts, QuotaFulls, LiveOnSite, IncidenceRate, QFIR) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (list[0], list[1], list[2], list[3], list[4], list[5], list[6], list[7], list[8], list[9], list[10],
             list[11], list[12], list[13]))
        conn.commit()  # saves to DB. Don't want to close the connection til I'm done using SQL in the program as open/closing wastes resources

    logging.debug('Now calling SQLite fn create_table')
    create_table()  # run the function above

    logging.debug('Now calling SQLite fn dynamic_data_entry')
    for list in list_of_projects:
        dynamic_data_entry(list)  # run the function above

    c.close()
    conn.close()


def send_email(user, pwd, recipient, subject, body):

    gmail_user = user
    gmail_pwd = pwd
    FROM = user
    TO = recipient if type(recipient) is list else [recipient]
    SUBJECT = subject
    TEXT = body

    # Prepare actual message
    message = """From: %s\nTo: %s\nSubject: %s\n\n%s
    """ % (FROM, ", ".join(TO), SUBJECT, TEXT)
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.ehlo()
        server.starttls()
        server.login(gmail_user, gmail_pwd)
        server.sendmail(FROM, TO, message)
        server.close()
        print('successfully sent the mail')
    except:
        print('failed to send mail')


def email_body_content(list_of_newbies):
    logging.debug('Initialising email_body_content function')
    body = ''
    for projectNumber in list_of_newbies:
        for project in latest10:
            if projectNumber in project:
                #print('Project found:',project)
                body += 'New project added to Admin. Project number: {} ; Project name: {}, Client name: {} \n\n'.format(project[3],project[1],project[4])
    #print(body)
    return(body)


def dict_creator(value_list):   # this function takes in a MO from the regex and creates and returns a per-project dict, with keys as per the headings below
    headings = ['URL', 'Survey name', 'Alias', 'Project number', 'Client name', 'junk', 'Expected LOI', 'Actual LOI', 'Completes', 'Screen Outs', 'Quota Fulls', 'Live on site']
    new_dict = {}
    for i in range(0, len(headings)):
        new_dict.setdefault(headings[i], value_list[i])
    completes = int(value_list[8])
    quota_fulls = int(value_list[10])
    screen_outs = int(value_list[9])
    if completes == 0:
        incidence = 0
        qf_incidence = 0
    else:
        try:
            incidence = (completes / (completes + screen_outs))
        except Exception as err:
            #print ('an exception occured: ', err)
            incidence = 0
        try:
            qf_incidence = (completes / (completes + screen_outs + quota_fulls))
        except Exception as err2:
            # print('an exception occured:',err2)
            qf_incidence = 0
    new_dict.setdefault('incidence', incidence)
    new_dict.setdefault('QFincidence', qf_incidence)
    return new_dict


def create_masterDict(mo):     #creates a dict of all project dicts in given MO
    master_dict = {}
    for i in range(0,len(mo)):
        url_aka_guid = mo[i][0]
        master_dict.setdefault(url_aka_guid, dict_creator(mo[i]))

    return master_dict


def excel_export_dict(dic, filename): # works with the 14 headings as per scrape (not the 22 mergedDict headings)
    logging.debug('Excel section - creating workbook object')
    wb = openpyxl.Workbook()  # create excel workbook object
    wb.save(filename)  # save workbook
    sheet = wb.active  # create sheet object as the Active sheet from the workbook object
    wb.save(filename)  # save workbook
    headings_list = ['URL','Survey name','Alias','Project number','Client name','junk','Expected LOI','Actual LOI','Completes','Screen Outs','Quota Fulls','Live on site', 'incidence', 'QFincidence']

    # this bit populates and emboldens the first row
    row = 1
    for column in range(0, len(headings_list)):
        cell = sheet.cell(row=row, column=column+1)
        cell.value = headings_list[column]
    make_bold(sheet, wb, sheet['A1':'O1'])    # Calls the make_bold function on first row of excel sheet

    # this bit then populates the rest of the sheet with the masterDict content
    for row, item_tuple in enumerate(dic.items(), 2):
        # print(f'row is {row}, item_tuple is {item_tuple}, key is {item_tuple[0]}, project dic is{item_tuple[1]}')
        for column, heading in enumerate(headings_list, 1):
            # print(f"row is {row}, column is {column} heading is {heading}, nested value is {item_tuple[1].get(heading)}")
            cell = sheet.cell(row=row, column=column)  # so on first loop, row = 2, col = 1
            v = item_tuple[1].get(heading)
            try:
                v = float(v)  # try to convert value to a float, so it will store numbers as numbers and not strings
            except ValueError:
                pass  # if it's not a number and therefore returns an error, don't try to convert it to a number
            cell.value = v
            if (column == 13) | (column == 14):  # for all cells in column 13 or 14 (IR / QFIR)
                cell.style = 'Percent'  # ... change cell format (style) to 'Percent', a built-in style within openpyxl

    wb.save(filename)  # save workbook as admin.xlsx
    logging.debug('Excel workbook completed and saved')


def mapping_dict_creator(excel_filename, r1, r2, c1, c2): # given excel filename and 2-column-wide excel table co-ordinates, creates a dictionary converting the table into key-value pairs
    logging.debug('Now attempting to read-in excel data to create dic')
    wb = openpyxl.load_workbook(excel_filename)
    sheet = wb.active
    dic = {}
    for row in range(r1,r2):
        for column in range(c1,c2):
            cell = sheet.cell(row = row, column = column)
            v = cell.value
            #print(f'row is {row}. Column is {column}, value is {v}')
            if column == c1:
                key = v
            else:
                value = v
        dic.setdefault(key, value)
    return dic


def create_merged_dict_with_old_data(old_data_dict, old_data_mapping_dict):
    merged = {}
    for k, v in old_data_dict.items():
        nested_dict = {}  # blank dict which we will add to merged_dict at the end of each loop
        for nk, nv in v.items():
            # print(nk, nv)
            equiv = old_data_mapping_dict.get(nk)
            if equiv != nk:
                # print(f'project {k} has {nk} re-assigned as {equiv} equal to {nv}')
                nested_dict.setdefault(equiv, nv)
            else:
                # print(f'project {k} has {nk} same as {equiv} so no re-assignment; equal to {nv}')
                nested_dict.setdefault(nk, nv)
        merged.setdefault(k, nested_dict)
    return merged


def add_new_data(new_data_dict, merged_data_dict, new_data_mapping_dict, T3_or_T2):
    for k, v in new_data_dict.items():
        if k not in merged_data_dict.keys():   # if a totally new project
            # print(f'{k} not found in merged so must be new')
            nested_dict = {}  # blank dict which we will add to merged_dict at the end of each loop
            for nk, nv in v.items():    # loop through the keys and values of the project
                # print(f'{k} has key {nk}, value {nv}')
                equiv = new_data_mapping_dict.get(nk)
                # print(f'equiv is {equiv}')
                nested_dict.setdefault(equiv, nv)
                nested_dict.setdefault('Completes_T1', 0)  # setting to blank as old data doesn't exist
                nested_dict.setdefault('Screen Outs_T1', 0) # setting to blank as old data doesn't exist
                nested_dict.setdefault('Quota Fulls_T1', 0) # setting to blank as old data doesn't exist
                if T3_or_T2 == "T3":
                    nested_dict.setdefault('Completes_T2', 0)  # setting to blank as old data doesn't exist
                    nested_dict.setdefault('Screen Outs_T2', 0) # setting to blank as old data doesn't exist
                    nested_dict.setdefault('Quota Fulls_T2', 0) # setting to blank as old data doesn't exist
            merged_data_dict.setdefault(k, nested_dict)

        else:
            # print(f'{k} found in merged_dict.keys, attempting to add to it')
            for nk, nv in v.items():    # loop through the keys and values of the project
                # print(nk, nv)
                equiv = new_data_mapping_dict.get(nk)
                if equiv not in merged_data_dict[k].keys():
                    # print(f'adding to {k}: {equiv} = {nv}')
                    merged_data_dict[k][equiv] = nv


def dynamic_field_adder(dic, T3_or_T2):  # add the dynamic fields (gaps, overnight) to merged_dict
    if T3_or_T2 == "T2":
        T_new = "T2"
        T_old = "T1"
    elif T3_or_T2 == "T3":
        T_new = "T3"
        T_old = "T2"
    for k, v in dic.items():
        c_gap = int(v[f'Completes_{T_new}']) - int(v[f'Completes_{T_old}'])
        v['Completes_gap'] = c_gap
        # print(f'Completes Gap for {k} is {c_gap}')
        s_gap = int(v[f'Screen Outs_{T_new}']) - int(v[f'Screen Outs_{T_old}'])
        v['Screen Outs_gap'] = s_gap
        # print(f'Screen Outs Gap for {k} is {s_gap}')
        q_gap = int(v[f'Quota Fulls_{T_new}']) - int(v[f'Quota Fulls_{T_old}'])
        v['Quota Fulls_gap'] = q_gap
        # print(f'Quota Fulls Gap for {k} is {q_gap}')
        try:
            oIR = (c_gap / (c_gap + s_gap))
            v['incidence_overnight'] = oIR
        except Exception as err:
            #print ('an exception occured: ', err)
            oIR = 0
            v['incidence_overnight'] = oIR
        try:
            oQFIR = (c_gap / (c_gap + s_gap + q_gap))
            v['QFincidence_overnight'] = oQFIR
        except Exception as err:
            #print ('an exception occured: ', err)
            oQFIR = 0
            v['QFincidence_overnight'] = oQFIR


def excel_export_mergedDict(dict, filename, headings):     #export merged dict to excel
    logging.debug('Attempting to export merged_dict to excel')
    wb = openpyxl.Workbook()  # create excel workbook object
    wb.save(filename)  # save workbook
    sheet = wb.active  # create sheet object as the Active sheet from the workbook object
    wb.save(filename)  # save workbook

    # this bit populates and emboldens the first row
    row = 1
    for column in range(0,len(headings)):
        cell = sheet.cell(row=row, column=column+1)
        cell.value = headings[column]
    make_bold(sheet, wb, sheet['A1':'Z1'])    #Calls the make_bold function on first row of excel sheet

    percentage_headings = ['incidence', 'incidence_overnight', 'QFincidence', 'QFincidence_overnight',]

     # this bit then populates the rest of the sheet with the merged_dict content
    for row, item_tuple in enumerate(dict.items(), 2):
        for column, heading in enumerate(headings, 1):
            cell = sheet.cell(row=row, column=column)  # so on first loop, row = 2, col = 1
            v = item_tuple[1].get(heading)
            try:
                v = float(v)  # try to convert value to a float, so it will store numbers as numbers and not strings
            except ValueError:
                pass  # if it's not a number and therefore returns an error, don't try to convert it to a number
            except TypeError:
                pass
            cell.value = v
            if heading in percentage_headings:  # for all cells with headings that should have % data
                cell.style = 'Percent'  # ... change cell format (style) to 'Percent', a built-in style within openpyxl
            if heading == 'Completes_gap':
                light_blue = 'A9CCE3'
                cell.fill = PatternFill("solid", fgColor=light_blue)
            if (heading == 'Screen Outs_gap') | (heading == 'Quota Fulls_gap'):
                orange = 'F8C471'
                cell.fill = PatternFill("solid", fgColor=orange)
    wb.save(filename)  # save workbook with given filename
    logging.debug('Excel workbook completed and saved')


def changes_dict_creator(large_dict):
    my_dict = {}
    for k, v in large_dict.items():
        if v['Completes_gap'] > 0 or v['Screen Outs_gap'] > 0 or v['Quota Fulls_gap'] > 0:
            my_dict.setdefault(k, v)
    return my_dict


def old_data_excel_to_dict_importer(excel_filename):  # given excel filename, creates a dictionary converting the table into key-value pairs
    logging.debug('Old data import - now attempting to read-in excel data to create dic')
    wb = openpyxl.load_workbook(excel_filename)
    sheet = wb.active
    dic = {}
    headers_dict = excel_headings_grabber(excel_filename)
    num_of_cols = column_counter(excel_filename)
    num_of_rows = row_counter(excel_filename)
    # print(f'#cols is {num_of_cols} and #rows is {num_of_rows}')
    for k, v in headers_dict.items():
        if v == 'URL':
            url_column = k
            # logging.debug(f'URL is in column {url_column}')
    for row in range(2,num_of_rows+1):
        nested_dict = {}
        for column in range(1,num_of_cols+1):
            cell = sheet.cell(row = row, column = column)
            v = cell.value
            # print(f'row is {row}. Column is {column}, value is {v}')
            nested_key = headers_dict[column]
            nested_val = v
            nested_dict.setdefault(nested_key, nested_val)
            if column == url_column:
                url = v
        dic.setdefault(url, nested_dict)
    return dic


def column_counter(xls_filename): #checks row 1 and counts how many cells have data, therefore how many columns in xls
    logging.debug('Counting columns in xls')
    wb = openpyxl.load_workbook(xls_filename)
    sheet = wb.active
    cols = 1 # start from 1st column
    while 1:
        cell = sheet.cell(row = 1, column = cols)
        v = cell.value
        if v != None: # if there is data in the cell
            cols += 1 # check the next column along
        else:    # if no data in the cell, then that's the last column, so break
            break
    c = int(cols)-1  # need to be minus one because it increments cols, then realises it's an empty cell
    logging.debug(f'# cols = {c}')
    return c


def row_counter(xls_filename): #checks column 1 and counts how many cells have data, therefore how many rows in xls
    logging.debug('Counting rows in xls')
    wb = openpyxl.load_workbook(xls_filename)
    sheet = wb.active
    rows = 1  # start from 1st column
    while 1:
        check_cell = sheet.cell(row = rows, column = 1)
        v = check_cell.value
        if v != None:  #if there is data in the cell
            rows += 1  # check the next column along
        else:    # if no data in the cell, then that's the last row, so break
             break
    r = int(rows)-1 # need to be minus one because it increments rows, then realises it's an empty cell
    logging.debug(f'#rows = {r}')
    return r


def excel_headings_grabber(xls_filename): # checks row 1 of xls and returns a dictionary showing col# & heading
    logging.debug('excel_headings_grabber - establishing headings/columns dict')
    wb = openpyxl.load_workbook(xls_filename)
    sheet = wb.active
    cols = 1
    dic = {}  # start from 1st column
    while 1:
        check_cell = sheet.cell(row = 1, column = cols)
        v = check_cell.value
        if v != None:  # if there is data in the cell
            dic.setdefault(cols, v)
            cols += 1 # check the next column along
        else:    # if no data in the cell, then that's the last column, so break
            break
    return dic


def create_stripped_dict(merged_data_dict, strip_mapping_dict):
    stripped = {}
    for k, v in merged_data_dict.items():
        nested_dict = {}  # blank dict which we will add to stripped_dict at the end of each loop
        for nk, nv in v.items():
            # print(nk, nv)
            if nk in strip_mapping_dict.keys():
                equiv = strip_mapping_dict.get(nk)
                # print(f'project {k} has {nk} re-assigned as {equiv} equal to {nv}')
                nested_dict.setdefault(equiv, nv)
        stripped.setdefault(k, nested_dict)
    return stripped






################################################
# TEST MODE - THIS IS WHERE EVERYTHING GETS CREATED

mo_T1 = process_soup(T1_soup, 'export/T1_string.txt', old_site_regex)   #parameter: newSoup or T1_soup for testing
T1_dict = create_masterDict(mo_T1)
excel_export_dict(T1_dict, 'export/T1.xlsx')
len_of_mo_T1 = len(mo_T1)
len_of_T1_dict = len(T1_dict)
rows_in_T1_xls = row_counter('export/T1.xlsx')
print(f'len of mo_T1 is {len_of_mo_T1} T1_dict is {len_of_T1_dict} whereas excel file has {rows_in_T1_xls} rows.')


mo_T2 = process_soup(T2_soup, 'export/T2_string.txt', old_site_regex)
T2_dict = create_masterDict(mo_T2)
excel_export_dict(T2_dict, 'export/T2.xlsx')
len_of_mo_T2 = len(mo_T2)
len_of_T2_dict = len(T2_dict)
rows_in_T2_xls = row_counter('export/T2.xlsx')
print(f'len of mo_T2 is {len_of_mo_T2} T2_dict is {len_of_T2_dict} whereas excel file has {rows_in_T2_xls} rows.')


mo_T3 = process_soup(T3_soup, 'export/T3_string.txt', old_site_regex)
T3_dict = create_masterDict(mo_T3)
excel_export_dict(T3_dict, 'export/T3.xlsx')
len_of_mo_T3 = len(mo_T3)
len_of_T3_dict = len(T3_dict)
rows_in_T3_xls = row_counter('export/T3.xlsx')
print(f'len of mo_T3 is {len_of_mo_T3} T3_dict is {len_of_T3_dict} whereas excel file has {rows_in_T3_xls} rows.')


# now to create merged dict

# first I need MAPPING DICTS: dictionaries which indicate which variable in the old/new data dictionaries respectively
# should be mapped to which variable in the merged dict
T1_map = mapping_dict_creator('public/mapping.xlsx', 3, 17, 1, 3)
T2_map = mapping_dict_creator('public/mapping.xlsx', 3, 17, 4, 6)
T3_map = mapping_dict_creator('public/mapping.xlsx', 3, 17, 7, 9)
strip_map = mapping_dict_creator('public/mapping.xlsx', 3, 17, 10, 12)

merged_dict = create_merged_dict_with_old_data(T1_dict, T1_map)
# now add all the new data, bearing in mind that the project may or may not already exist in merged_dict
add_new_data(T2_dict, merged_dict, T2_map, "T2")
# now let's add the formula-calculated fields within each dict
merged_dict_headings_2_data_sets = ['URL',
'Survey name',
'Alias',
'Project number',
'Client name',
'junk',
'Expected LOI',
'Actual LOI',
'Completes_T1',
'Completes_T2',
'Completes_gap',
'Screen Outs_T1',
'Screen Outs_T2',
'Screen Outs_gap',
'Quota Fulls_T1',
'Quota Fulls_T2',
'Quota Fulls_gap',
'Live on site',
'incidence',
'incidence_overnight',
'QFincidence',
'QFincidence_overnight']
merged_dict_headings_3_data_sets = ['URL',
'Survey name',
'Alias',
'Project number',
'Client name',
'junk',
'Expected LOI',
'Actual LOI',
'Completes_T1',
'Completes_T2',
'Completes_T3',
'Completes_gap',
'Screen Outs_T1',
'Screen Outs_T2',
'Screen Outs_T3',
'Screen Outs_gap',
'Quota Fulls_T1',
'Quota Fulls_T2',
'Quota Fulls_T3',
'Quota Fulls_gap',
'Live on site',
'incidence',
'incidence_overnight',
'QFincidence',
'QFincidence_overnight']

add_new_data(T3_dict, merged_dict, T3_map, "T3")  # add T3 data to merged_dict
dynamic_field_adder(merged_dict, "T3")  # add the dynamic fields (gaps, overnight) to merged_dict, assuming T3 is latest data
excel_export_mergedDict(merged_dict, 'export/merged.xlsx', merged_dict_headings_2_data_sets) # excel export of merged_dict
# excel_export_mergedDict(merged_dict, 'export/merged.xlsx', merged_dict_headings_3_data_sets) # excel export of merged_dict
len_of_merged_dict = len(merged_dict)
rows_in_merged_xls = row_counter('export/merged.xlsx')
print(f'len of merged_dict is {len_of_merged_dict} whereas excel file has {rows_in_merged_xls} rows.')


# If Comp, SO or QF gaps > 0, then project has changed. Add it to a 'changed' dictionary, and export that to excel, excluding junk/alias/URL fields
changes_dict = changes_dict_creator(merged_dict)

# only certain headings are of interest in the new 'changes' excel export, they are in this list
changes_dict_headings_of_interest = [
'Survey name','Project number','Client name','Expected LOI','Actual LOI','Completes_T1','Completes_T2',
    'Completes_gap','Screen Outs_T1','Screen Outs_T2','Screen Outs_gap','Quota Fulls_T1',
    'Quota Fulls_T2','Quota Fulls_gap','incidence','incidence_overnight','QFincidence','QFincidence_overnight',
]
excel_export_mergedDict(changes_dict, 'export/changes_dict.xlsx', changes_dict_headings_of_interest)  # excel export of changes_dict using columns of interest only






################################################
# LIVE MODE - THIS IS WHERE EVERYTHING GETS CREATED

# STEPS IN LOOP
# 1 import old data from xls, store in dict as D1
# 2 download new data, store in dict as D2
# 3 create merged file
# 4 create report showing changes only
# 5 send email

# 1 import old data from xls, store in dict as D1

imported_dict = old_data_excel_to_dict_importer('export/merged_to_import.xlsx')
# create a dic which is only the stripped out fields of interest i.e. stripping back from a merged dict to unmerged
stripped_dict = create_stripped_dict(imported_dict, strip_map)

# 2 download new data, store in dict as D2

# here is the code to actually do the download:
"""
D2_soup = download_soup()     # toggle off for test mode
mo_D2 = process_soup(D2_soup, 'export/D2_string.txt', new_site_regex)   # parameter: D2_soup or T1_soup for testing, plus string txt filename
D2_dict = create_masterDict(mo_D2)
excel_export_dict(D2_dict, 'export/D2.xlsx')
len_of_mo_D2 = len(mo_D2)
len_of_D2_dict = len(D2_dict)
rows_in_D2_xls = row_counter('export/D2.xlsx')
print(f'len of mo_D2 is {len_of_mo_D2} D2_dict is {len_of_D2_dict} whereas excel file has {rows_in_D2_xls} rows.')
"""

# or to just use the pre-downloaded table string:
mo_D2_backup = process_string('export/D2_string_backup.txt', new_site_regex)
D2_backup_dict = create_masterDict(mo_T3)
excel_export_dict(D2_backup_dict, 'export/D2_backup.xlsx')
len_of_mo_D2_backup = len(mo_D2_backup)
len_of_D2_backup_dict = len(D2_backup_dict)
rows_in_D2_backup_xls = row_counter('export/D2_backup.xlsx')
print(f'len of mo_D2_backup is {len_of_mo_D2_backup} D2_backup_dict is {len_of_D2_backup_dict} whereas excel file has {rows_in_D2_backup_xls} rows.')


# 3 create merged/changes files
# This section below needs to be debugged carefully as it's copy/pasted from above.
# To test/debug, I should create a mergedDict where recent (T3) data is used then strip that so that changesDict is smaller
# and more comprehensible to test

D_merged_dict = create_merged_dict_with_old_data(stripped_dict, T1_map)
# now add all the new data, bearing in mind that the project may or may not already exist in merged_dict
add_new_data(D2_backup_dict, D_merged_dict, T2_map, "T2")
dynamic_field_adder(D_merged_dict, "T2")  # add the dynamic fields (gaps, overnight) to merged_dict, assuming T3 is latest data
excel_export_mergedDict(D_merged_dict, 'export/D_merged.xlsx', merged_dict_headings_2_data_sets) # excel export of merged_dict
len_of_D_merged_dict = len(D_merged_dict)
rows_in_D_merged_xls = row_counter('export/D_merged.xlsx')
print(f'len of D_merged_dict is {len_of_D_merged_dict} whereas excel file has {rows_in_D_merged_xls} rows.')


# If Comp, SO or QF gaps > 0, then project has changed. Add it to a 'changed' dictionary, and export that to excel, excluding junk/alias/URL fields
D_changes_dict = changes_dict_creator(D_merged_dict)

# only certain headings are of interest in the new 'changes' excel export, they are in this list
changes_dict_headings_of_interest = [
'Survey name','Project number','Client name','Expected LOI','Actual LOI','Completes_T1','Completes_T2',
    'Completes_gap','Screen Outs_T1','Screen Outs_T2','Screen Outs_gap','Quota Fulls_T1',
    'Quota Fulls_T2','Quota Fulls_gap','incidence','incidence_overnight','QFincidence','QFincidence_overnight',
]
excel_export_mergedDict(D_changes_dict, 'export/D_changes_dict.xlsx', changes_dict_headings_of_interest)  # excel export of changes_dict using columns of interest only








"""
original10 = create_top_list(mo_T1, 10)   #match object, desired number of projects in list
while 1:     #this is the loop that endlessly repeats
    #newSoup = download_soup()                # download latest HTML; toggle off for test mode
    mo2 = process_soup(T2_soup, 'mo2_string.txt', old_site_regex)   # parameter can be newSoup for live or T2_soup for test mode
    latest10 = create_top_list(mo2, 10)
    newbies = new_project_search(latest10,original10)   #parameters should be latest10 and original10
    print('Latest10 looks like this:\n',latest10)
    print('Original10 looks like this:\n',original10)
    print('newbies:\n',newbies)
    if len(newbies) > 0:
        send_email(cfg.my_gmail_uname, cfg.my_gmail_pw, cfg.my_work_email,'Admin: new project added',email_body_content(newbies))
    original10 = latest10    #overwrite original10 with the latest10
    print('End of program, waiting 60 sec')
    time.sleep(1000)     #1000 for test mode
"""










"""
original20 = create_top_list(mo_T1, 20)   #match object, desired number of projects in list
new20 = create_top_list(mo_T2, 20)
print("new projects are: ", new_project_search(new20, original20))
print("original20 looks like this: ",original20)
print("the job# in the first item in original20 looks like this: ", original20[0][3])
"""


# for all job #s in new20, if job# appears in original20:
    # for all non-numerical items, their equivalent in changed20 is identical to new20
    # for all numerical items in that job for new20:
        # changed20 equivalent = new20 number minus original20 number






"""

#TO DO: compare original10 and latest10 and flag any 'zero to 1' completes movement(new function)


#masterList = create_masterList(mo2)   #match object
#excel_export(latest10)           #list to export
#export_to_sqlite(original10)       #list to export



#This is a test sequence, to compare lists generated from old and new soup
#It works beautifully when I'm looking for 10 and 20 list length, but for 30 I get an error. Not sure why a project was being searched for and attempted removal twice, but added 'try and except' logic to keep program running


# logging.debug('Example sequence')
# exampleOldMo = process_soup(T1_soup, 'mo_old_string.txt', old_site_regex)
# exampleNewMo = process_soup(T2_soup, 'mo_new_string.txt', old_site_regex)
# exampleOriginal10 = create_top_list(exampleOldMo, 10)   #match object, desired number of projects in list
# print('ExampleOriginal10:\n', exampleOriginal10,'\n')
# exampleLatest10 = create_top_list(exampleNewMo, 10)   #match object, desired number of projects in list
# print('ExampleLatest10:\n', exampleLatest10,'\n')
# exampleNewbies = new_project_search(exampleLatest10, exampleOriginal10)
# print('The example new projects are:\n',exampleNewbies,'\n')
# send_email(cfg.my_gmail_uname, cfg.my_gmail_pw, cfg.my_work_email,'Admin: new project added',str(exampleNewbies))
"""

